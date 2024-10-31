# -*- coding: utf-8 -*-
"""
Created on Thu Oct 31 09:27:30 2024

@author: user
"""

from flask import Flask, render_template
import csv
import os
import glob

from openpyxl import Workbook
from openpyxl.styles import NamedStyle

def create_excel_style(wb, style_name, number_format):
    """
    在指定的活頁簿中建立或取得指定名稱的格式。

    參數:
    wb (Workbook): 要操作的活頁簿物件。
    style_name (str): 格式的名稱。
    number_format (str): 格式的數字格式。

    返回:
    NamedStyle: 新建立或既有的格式。
    """
    # 檢查格式是否已經存在
    if style_name in wb.named_styles:
        return wb.named_styles[style_name]
    
    # 建立新的樣式
    new_style = NamedStyle(name=style_name, number_format=number_format)
    wb.add_named_style(new_style)
    return new_style

def apply_excel_styles(wb, ws):
    """
    根據欄位名稱套用不同格式到工作表中的儲存格。

    參數:
    wb (Workbook): openpyxl 的 Workbook 物件。
    ws (Worksheet): openpyxl 的 Worksheet 物件。
    """
    # 建立不同類型的格式
    string_style = create_excel_style(wb, 'string_style', '@')
    date_style = create_excel_style(wb, 'date_style', 'YYYY/MM/DD')
    number_style = create_excel_style(wb, 'number_style', '#,##0;-#,##0')
    
    # 根據欄位名稱套用儲存格樣式
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if '日期' in header or '交期' in header:
                cell.style = date_style
            elif '數量' in header or '金額' in header:
                cell.style = number_style
            else:
                cell.style = string_style

def create_sample_excel() -> 'html':    
    """
    建立一個 demo 的 Excel 檔案並套用儲存格格式。
    """
    # 建立活頁簿與工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "demo"
    
    # 增加欄位名稱與資料
    headers = ['訂單編號', '數量', '單價', '總金額', '出貨日期']
    data = [
        ['0001', 100, 10.5, 1050, '2024/09/13'],
        ['0002', 200, 15.0, 3000, '2024/10/01'],
        ['0003', 50, 20.0, 1000, '2024/11/20']
    ]
    
    ws.append(headers)
    for row in data:
        ws.append(row)
    
    # 套用樣式
    apply_excel_styles(wb, ws)
    
    # 儲存 Excel 檔案
    wb.save('styled_example.xlsx')

    


def the_new_arrival(path) -> 'html':    
    contents_new = []
    files = glob.glob(os.path.join(path, "*.csv"))
    file_name = ''
    rCount_new = 0
    for file in files: 
        file_name = os.path.basename(file)
        if file_name == "newArrival.csv":
            sourceFile = open(file, "r",encoding="utf-8")
            sourceReader = csv.reader(sourceFile)
            for row in sourceReader:
                contents_new.append([])
                if sourceReader.line_num > 2:
                    contents_new[-1].append(row[1])
                    contents_new[-1].append(row[7])
                    contents_new[-1].append(row[5])
                    contents_new[-1].append(row[2])
                    rCount_new += 1
       
    titles = ('部門', '職稱', '姓名', '報到日')
    return render_template('viewlog.html',
                           the_title='New Arrival',
                           the_file=file_name,
                           row_count=rCount_new,
                           the_row_titles=titles,
                           the_data=contents_new,)